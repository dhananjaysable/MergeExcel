import os
import sys
import subprocess
import importlib.util

# --- Auto-Installation of Dependencies ---
def install_and_import(package, import_name=None):
    import_name = import_name or package
    if importlib.util.find_spec(import_name) is None:
        print(f"Installing {package}...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            print(f"Successfully installed {package}")
        except subprocess.CalledProcessError as e:
            print(f"Failed to install {package}. Error: {e}")
            sys.exit(1)

# List of required packages
required_packages = [
    ("openpyxl", "openpyxl"),
    ("PyQt6", "PyQt6"),
    ("PyQt6-Fluent-Widgets", "qfluentwidgets")
]

# Install missing packages
for package, import_name in required_packages:
    install_and_import(package, import_name)

# Force qfluentwidgets to use PyQt6 (Must be set before importing qfluentwidgets)
os.environ["QT_API"] = "pyqt6"

import threading
import pathlib
from copy import copy
from datetime import datetime
import webbrowser
import platform
import warnings

# Now safe to import
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSize
from PyQt6.QtWidgets import (
    QApplication, QFileDialog, QTableWidgetItem, QHeaderView, QFrame, 
    QVBoxLayout, QHBoxLayout, QWidget, QSizePolicy
)
from PyQt6.QtGui import QIcon, QColor, QFont

# Create QApplication BEFORE importing qfluentwidgets to avoid "Must construct a QApplication" error
# This is necessary because qfluentwidgets might initialize widgets at module level or during import
if QApplication.instance() is None:
    app = QApplication(sys.argv)
else:
    app = QApplication.instance()

from qfluentwidgets import (
    FluentWindow, SubtitleLabel, PrimaryPushButton, LineEdit, PushButton, 
    TableWidget, CheckBox, ProgressBar, TextEdit, 
    InfoBar, InfoBarPosition, Theme, setTheme, setThemeColor,
    StrongBodyLabel, CaptionLabel, BodyLabel, CardWidget,
    TransparentToolButton, FluentIcon as FIF,
    TitleLabel, ComboBox, SwitchButton
)

# Suppress warnings about data validation
warnings.filterwarnings('ignore', category=UserWarning)

# --- Data Structures ---


class ExcelFileInfo:
    """Stores metadata about an Excel file found in the scan."""
    def __init__(self, path, display_name):
        self.path = pathlib.Path(path)
        self.display_name = display_name
        self.sheet_names = []
        self.sheet_count = 0
        self.selected = True  # Default to checked


class MergeSettings:
    """Stores configuration for the merge operation."""
    def __init__(self):
        self.include_subfolders = False
        self.skip_temp_files = True
        self.output_folder = pathlib.Path("")
        self.output_filename = "MergedWorkbook.xlsx"
        self.create_index_sheet = False
        self.preserve_formulas = True


# --- Core Logic Classes ---


class FolderScanner:
    """Responsible for finding Excel files and extracting metadata."""

    @staticmethod
    def scan(folder_path, include_subfolders=False, skip_temp=True):
        folder = pathlib.Path(folder_path)
        if not folder.exists():
            return []

        found_files = []
        pattern = "**/*.xls*" if include_subfolders else "*.xls*"

        for file_path in folder.glob(pattern):
            suffix = file_path.suffix.lower()
            if suffix not in (".xlsx", ".xlsm"):
                continue

            if skip_temp and file_path.name.startswith("~$"):
                continue

            info = ExcelFileInfo(file_path, file_path.name)

            try:
                wb = load_workbook(
                    file_path,
                    read_only=True,
                    keep_links=False,
                    data_only=False,
                )
                info.sheet_names = wb.sheetnames
                info.sheet_count = len(info.sheet_names)
                wb.close()
                found_files.append(info)
            except Exception as e:
                print(f"Skipping {file_path.name}: {e}")

        found_files.sort(key=lambda x: str(x.path).lower())
        return found_files


class EnhancedSheetCopier:
    """
    Simple & safe helper to copy content & style from one sheet to another.

    - Copies values + styles
    - Copies merged cells
    - Copies column widths & row heights
    - Copies freeze panes
    - DOES NOT copy tables, CF, charts, defined names, etc.
      (this avoids all Excel XML corruption issues)
    """

    @staticmethod
    def copy_sheet(source_ws, target_wb, new_title, preserve_formulas=True):
        # Excel sheet name max 31 chars, no :\\/?*[]
        safe_title = (
            new_title
            .replace(":", "_")
            .replace("/", "_")
            .replace("\\", "_")
            .replace("?", "_")
            .replace("*", "_")
            .replace("[", "_")
            .replace("]", "_")
        )[:31]

        base = safe_title or "Sheet"
        final = base
        c = 1
        while final in target_wb.sheetnames:
            final = (base[:28] + "_" + str(c))[:31]
            c += 1

        target_ws = target_wb.create_sheet(final)

        try:
            # 1. Sheet properties
            try:
                if source_ws.sheet_properties.tabColor:
                    target_ws.sheet_properties.tabColor = copy(source_ws.sheet_properties.tabColor)
            except Exception:
                pass

            # 2. Sheet state (visible, hidden, very hidden)
            try:
                target_ws.sheet_state = source_ws.sheet_state
            except Exception:
                pass

            # 3. Freeze panes
            try:
                if source_ws.freeze_panes:
                    target_ws.freeze_panes = source_ws.freeze_panes
            except Exception:
                pass

            # 4. Print settings and page setup
            try:
                if source_ws.page_setup:
                    target_ws.page_setup.orientation = source_ws.page_setup.orientation
                    target_ws.page_setup.paperSize = source_ws.page_setup.paperSize
                    target_ws.page_setup.fitToPage = source_ws.page_setup.fitToPage
                    target_ws.page_setup.fitToHeight = source_ws.page_setup.fitToHeight
                    target_ws.page_setup.fitToWidth = source_ws.page_setup.fitToWidth
            except Exception:
                pass

            # 5. Print options
            try:
                if source_ws.print_options:
                    target_ws.print_options.horizontalCentered = source_ws.print_options.horizontalCentered
                    target_ws.print_options.verticalCentered = source_ws.print_options.verticalCentered
            except Exception:
                pass

            # 6. Column widths
            for col_letter, col_dim in source_ws.column_dimensions.items():
                try:
                    td = target_ws.column_dimensions[col_letter]
                    if col_dim.width:
                        td.width = col_dim.width
                    td.hidden = col_dim.hidden
                except Exception:
                    pass

            # 7. Row heights
            for r, rd_src in source_ws.row_dimensions.items():
                try:
                    rd_tgt = target_ws.row_dimensions[r]
                    if rd_src.height:
                        rd_tgt.height = rd_src.height
                    rd_tgt.hidden = rd_src.hidden
                except Exception:
                    pass

            # 8. Collect merged cells (apply after cell copying)
            merged_ranges = []
            for merged in source_ws.merged_cells.ranges:
                try:
                    merged_ranges.append(str(merged))
                except Exception:
                    pass
            
            # 9. Cell values + styles (Direct Copy)
            for row in source_ws.iter_rows():
                for cell in row:
                    try:
                        # Get target cell
                        target_cell = target_ws.cell(row=cell.row, column=cell.column)
                        
                        # Copy value or formula
                        if preserve_formulas and cell.data_type == 'f':
                            try:
                                target_cell.value = cell.value
                            except Exception:
                                target_cell.value = cell.value if cell.value else ""
                        else:
                            target_cell.value = cell.value
                        
                        # Copy formatting
                        try:
                            if cell.font: target_cell.font = copy(cell.font)
                        except: pass
                        
                        try:
                            if cell.border: target_cell.border = copy(cell.border)
                        except: pass
                        
                        try:
                            if cell.fill: target_cell.fill = copy(cell.fill)
                        except: pass
                        
                        try:
                            if cell.number_format: target_cell.number_format = cell.number_format
                        except: pass
                        
                        try:
                            if cell.alignment: target_cell.alignment = copy(cell.alignment)
                        except: pass
                        
                        try:
                            if cell.protection: target_cell.protection = copy(cell.protection)
                        except: pass
                    
                    except Exception as e:
                        continue

            # 10. Apply merged cells AFTER all cells are copied
            for merged_range in merged_ranges:
                try:
                    target_ws.merge_cells(merged_range)
                except Exception:
                    pass

            # 11. Copy data validations
            try:
                if hasattr(source_ws, 'data_validations') and source_ws.data_validations:
                    for dv in source_ws.data_validations.dataValidation:
                        try:
                            target_ws.add_data_validation(copy(dv))
                        except Exception:
                            pass
            except Exception:
                pass

            # 12. Copy conditional formatting AND Differential Styles (DXF)
            try:
                if hasattr(source_ws, 'conditional_formatting') and source_ws.conditional_formatting:
                    # Access source workbook to get DXF styles
                    source_wb = source_ws.parent
                    
                    for cf_range, cf_rules in source_ws.conditional_formatting._cf_rules.items():
                        try:
                            for rule in cf_rules:
                                # Copy the rule
                                new_rule = copy(rule)
                                
                                # Handle DXF (Differential Style) copying
                                if hasattr(new_rule, 'dxfId') and new_rule.dxfId is not None:
                                    try:
                                        # Get source DXF style
                                        if hasattr(source_wb, '_differential_styles') and len(source_wb._differential_styles) > new_rule.dxfId:
                                            source_dxf = source_wb._differential_styles[new_rule.dxfId]
                                            
                                            # Add to target workbook if not present
                                            if not hasattr(target_wb, '_differential_styles'):
                                                target_wb._differential_styles = []
                                            
                                            # Check if this style already exists in target to reuse ID
                                            # (Simple check might not be enough, but appending is safer for now)
                                            target_wb._differential_styles.append(copy(source_dxf))
                                            new_dxf_id = len(target_wb._differential_styles) - 1
                                            
                                            # Update rule to point to new DXF ID
                                            new_rule.dxfId = new_dxf_id
                                    except Exception as e:
                                        pass

                                target_ws.conditional_formatting.add(cf_range, new_rule)
                        except Exception:
                            pass
            except Exception:
                pass

            # 13. Copy Excel Tables (Native Object Copying)
            try:
                if hasattr(source_ws, 'tables') and source_ws.tables:
                    from openpyxl.worksheet.table import Table, TableStyleInfo
                    
                    for table_name in source_ws.tables:
                        try:
                            source_table = source_ws.tables[table_name]
                            
                            # Create unique table name for target
                            base_table_name = f"{final}_{source_table.displayName}"
                            # Sanitize table name (no spaces, special chars)
                            safe_table_name = (
                                base_table_name
                                .replace(" ", "_")
                                .replace("-", "_")
                                .replace(".", "_")
                                .replace(":", "_")
                                .replace("/", "_")
                                .replace("\\", "_")
                            )[:255]
                            
                            # Ensure unique table name
                            final_table_name = safe_table_name
                            counter = 1
                            while final_table_name in target_ws.tables:
                                final_table_name = f"{safe_table_name}_{counter}"
                                counter += 1
                            
                            # Create new table with same range
                            new_table = Table(
                                displayName=final_table_name,
                                ref=source_table.ref
                            )
                            
                            # Copy table style info
                            if hasattr(source_table, 'tableStyleInfo') and source_table.tableStyleInfo:
                                new_table.tableStyleInfo = TableStyleInfo(
                                    name=source_table.tableStyleInfo.name,
                                    showFirstColumn=source_table.tableStyleInfo.showFirstColumn,
                                    showLastColumn=source_table.tableStyleInfo.showLastColumn,
                                    showRowStripes=source_table.tableStyleInfo.showRowStripes,
                                    showColumnStripes=source_table.tableStyleInfo.showColumnStripes
                                )
                            
                            # Add table to target worksheet
                            target_ws.add_table(new_table)
                            
                        except Exception as e:
                            print(f"  Warning: Could not copy table '{table_name}': {e}")
                            pass
            except Exception as e:
                print(f"Warning: Error copying tables: {e}")
                pass

            return target_ws

        except Exception as e:
            print(f"Copy error: {e}")
            import traceback
            traceback.print_exc()
            return target_ws



class ExcelMerger:
    """Orchestrator for the merge process (openpyxl-only)."""

    @staticmethod
    def _build_sheet_name(file_index: int, sheet_name: str, existing_names) -> str:
        """Build a collision-safe sheet name with index prefix within 31-char limit."""
        safe_sheet = (
            sheet_name
            .replace(":", "_")
            .replace("/", "_")
            .replace("\\", "_")
            .replace("?", "_")
            .replace("*", "_")
            .replace("[", "_")
            .replace("]", "_")
        )
        base = f"{file_index}_{safe_sheet}"
        if len(base) > 31:
            prefix = f"{file_index}_"
            remaining = max(31 - len(prefix), 1)
            base = prefix + safe_sheet[:remaining]

        name = base
        suffix = 1
        existing = set(existing_names)
        while name in existing:
            candidate = f"{base}_{suffix}"
            if len(candidate) > 31:
                candidate = candidate[:31]
            name = candidate
            suffix += 1
        return name

    @staticmethod
    def merge(files, settings, log_cb, progress_cb):
        try:
            files_to_process = [f for f in files if f.selected]
            if not files_to_process:
                raise ValueError("No files selected for merging.")

            total_sheets = sum(f.sheet_count for f in files_to_process)
            current_sheet_count = 0

            target_wb = openpyxl.Workbook()
            # remove default sheet
            if target_wb.active:
                target_wb.remove(target_wb.active)

            mapping_data = []

            log_cb("Initializing merge process (openpyxl)...")
            log_cb(f"Preserving formulas: {settings.preserve_formulas}")

            for file_idx, file_info in enumerate(files_to_process, start=1):
                log_cb(f"Processing File {file_idx}/{len(files_to_process)}: {file_info.display_name}")

                try:
                    source_wb = load_workbook(
                        file_info.path,
                        data_only=not settings.preserve_formulas,
                        keep_links=False,
                        keep_vba=False,
                    )
                except Exception as e:
                    log_cb(f"ERROR opening file {file_info.display_name}: {e}")
                    continue

                try:
                    for sheet_name in source_wb.sheetnames:
                        try:
                            source_ws = source_wb[sheet_name]

                            new_sheet_name = ExcelMerger._build_sheet_name(
                                file_idx, sheet_name, target_wb.sheetnames
                            )
                            log_cb(f"  > Copying '{sheet_name}' -> '{new_sheet_name}'")

                            EnhancedSheetCopier.copy_sheet(
                                source_ws,
                                target_wb,
                                new_sheet_name,
                                preserve_formulas=settings.preserve_formulas,
                            )

                            mapping_data.append(
                                {
                                    "File Index": file_idx,
                                    "File Name": file_info.display_name,
                                    "Original Sheet": sheet_name,
                                    "New Sheet": new_sheet_name,
                                }
                            )

                            current_sheet_count += 1
                            progress_cb(current_sheet_count, total_sheets)
                        except Exception as e:
                            log_cb(f"ERROR copying sheet '{sheet_name}': {e}")
                            import traceback
                            log_cb(f"Traceback: {traceback.format_exc()}")
                            continue

                finally:
                    try:
                        source_wb.close()
                    except Exception:
                        pass

            # ---- Index sheet ----
            if settings.create_index_sheet and mapping_data:
                try:
                    log_cb("Generating Index sheet...")
                    index_ws = target_wb.create_sheet("Index", 0)

                    title_cell = index_ws.cell(row=1, column=1, value="Merged Workbook Index")
                    title_cell.font = openpyxl.styles.Font(bold=True, size=14)
                    index_ws.merge_cells("A1:D1")

                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    index_ws.cell(row=2, column=1, value=f"Generated on: {timestamp}")

                    headers = ["File Index", "File Name", "Original Sheet", "New Sheet"]
                    header_fill = openpyxl.styles.PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                    for col_idx, header in enumerate(headers, start=1):
                        cell = index_ws.cell(row=4, column=col_idx, value=header)
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = header_fill

                    for row_idx, row_data in enumerate(mapping_data, start=5):
                        index_ws.cell(row=row_idx, column=1, value=row_data.get("File Index", ""))
                        index_ws.cell(row=row_idx, column=2, value=row_data.get("File Name", ""))
                        index_ws.cell(row=row_idx, column=3, value=row_data.get("Original Sheet", ""))
                        index_ws.cell(row=row_idx, column=4, value=row_data.get("New Sheet", ""))

                    # simple hyperlinks to sheets
                    for row_idx in range(5, 5 + len(mapping_data)):
                        sheet_name = index_ws.cell(row=row_idx, column=4).value
                        if sheet_name and sheet_name in target_wb.sheetnames:
                            cell = index_ws.cell(row=row_idx, column=4)
                            cell.hyperlink = f"#'{sheet_name}'!A1"
                            cell.font = openpyxl.styles.Font(color="0563C1", underline="single")

                    # Auto column widths
                    try:
                        for col_idx in range(1, 5):  # Columns A-D
                            max_length = 0
                            for row_idx in range(1, 5 + len(mapping_data)):
                                try:
                                    cell_value = index_ws.cell(row=row_idx, column=col_idx).value
                                    if cell_value is not None:
                                        max_length = max(max_length, len(str(cell_value)))
                                except Exception:
                                    pass
                            
                            column_letter = get_column_letter(col_idx)
                            index_ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                    except Exception as e:
                        log_cb(f"Warning: Could not auto-size columns: {e}")

                    index_ws.freeze_panes = "A5"
                except Exception as e:
                    log_cb(f"ERROR creating index sheet: {e}")
                    import traceback
                    log_cb(f"Traceback: {traceback.format_exc()}")

            output_full_path = settings.output_folder / settings.output_filename
            log_cb(f"Saving to {output_full_path}...")
            target_wb.save(output_full_path)
            target_wb.close()

            log_cb("✓ Merge Complete!")
            log_cb(f"✓ File saved: {output_full_path}")

            return output_full_path
        except Exception as e:
            log_cb(f"CRITICAL ERROR in merge: {e}")
            import traceback
            log_cb(f"Traceback: {traceback.format_exc()}")
            raise


# --- GUI Application ---

class MergeWorker(QThread):
    progress_signal = pyqtSignal(int, int)  # current, total
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(str) # output path
    error_signal = pyqtSignal(str)

    def __init__(self, files, settings):
        super().__init__()
        self.files = files
        self.settings = settings

    def run(self):
        try:
            def log_cb(msg):
                self.log_signal.emit(msg)
            
            def progress_cb(current, total):
                self.progress_signal.emit(current, total)

            output_path = ExcelMerger.merge(
                self.files, self.settings, log_cb, progress_cb
            )
            self.finished_signal.emit(str(output_path))
        except Exception as e:
            self.error_signal.emit(str(e))

class ExcelMergerWindow(FluentWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Advanced Excel Merger")
        self.resize(1100, 800)
        self.setWindowIcon(QIcon("icon.ico"))
        
        # Theme
        setTheme(Theme.LIGHT)
        setThemeColor('#0078D4')

        self.files_data = []
        self.current_source_folder = ""
        self.last_output_path = None

        self.init_ui()

    def init_ui(self):
        self.main_widget = QWidget()
        self.main_widget.setObjectName("mergeInterface")
        # self.setCentralWidget(self.main_widget) # Not available in FluentWindow
        
        # Add the main widget as a sub-interface
        # We give it a name and icon to show in the navigation bar (even if we have only one)
        self.addSubInterface(self.main_widget, FIF.HOME, "Merge")

        self.v_layout = QVBoxLayout(self.main_widget)
        self.v_layout.setContentsMargins(30, 30, 30, 30)
        self.v_layout.setSpacing(20)

        # Title
        self.title_label = TitleLabel("📊 Advanced Excel Merger", self.main_widget)
        self.subtitle_label = CaptionLabel("Preserves formulas, formatting, and tables", self.main_widget)
        self.subtitle_label.setTextColor(QColor(100, 100, 100), QColor(200, 200, 200))
        
        title_layout = QVBoxLayout()
        title_layout.addWidget(self.title_label)
        title_layout.addWidget(self.subtitle_label)
        self.v_layout.addLayout(title_layout)

        # Source Selection Card
        self.source_card = CardWidget(self)
        source_layout = QVBoxLayout(self.source_card)
        source_layout.setContentsMargins(20, 20, 20, 20)
        
        source_header = StrongBodyLabel("Source Folder", self.source_card)
        source_layout.addWidget(source_header)

        h_source = QHBoxLayout()
        self.source_path_edit = LineEdit(self.source_card)
        self.source_path_edit.setPlaceholderText("Select a folder containing Excel files...")
        self.source_path_edit.setReadOnly(True)
        
        self.btn_browse = PushButton("Browse", self.source_card, FIF.FOLDER)
        self.btn_browse.clicked.connect(self.browse_source)
        
        self.btn_scan = PrimaryPushButton("Scan Folder", self.source_card, FIF.SYNC)
        self.btn_scan.clicked.connect(self.scan_folder)

        h_source.addWidget(self.source_path_edit, 1)
        h_source.addWidget(self.btn_browse)
        h_source.addWidget(self.btn_scan)
        source_layout.addLayout(h_source)
        
        self.v_layout.addWidget(self.source_card)

        # File List
        self.table = TableWidget(self)
        self.table.setBorderVisible(True)
        self.table.setBorderRadius(8)
        self.table.setWordWrap(False)
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["✓", "File Name", "Sheets", "Size", "Path"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().hide()
        
        self.v_layout.addWidget(self.table, 1)

        # File Actions
        h_file_actions = QHBoxLayout()
        self.btn_select_all = PushButton("Select All", self, FIF.CHECKBOX)
        self.btn_select_all.clicked.connect(lambda: self.toggle_all(True))
        
        self.btn_deselect_all = PushButton("Deselect All", self, FIF.CANCEL)
        self.btn_deselect_all.clicked.connect(lambda: self.toggle_all(False))
        
        self.lbl_file_count = BodyLabel("0 files found", self)
        
        h_file_actions.addWidget(self.btn_select_all)
        h_file_actions.addWidget(self.btn_deselect_all)
        h_file_actions.addWidget(self.lbl_file_count)
        h_file_actions.addStretch()
        
        self.v_layout.addLayout(h_file_actions)

        # Settings & Output Card
        self.settings_card = CardWidget(self)
        settings_layout = QVBoxLayout(self.settings_card)
        settings_layout.setContentsMargins(20, 20, 20, 20)

        # Grid for settings
        h_settings = QHBoxLayout()
        
        # Left: Search & Options
        v_opts = QVBoxLayout()
        v_opts.addWidget(StrongBodyLabel("Options", self.settings_card))
        
        self.chk_subfolders = CheckBox("Include Subfolders", self.settings_card)
        self.chk_skip_temp = CheckBox("Skip Temporary Files (~$)", self.settings_card)
        self.chk_skip_temp.setChecked(True)
        self.chk_preserve = CheckBox("Preserve Formulas", self.settings_card)
        self.chk_preserve.setChecked(True)
        self.chk_index = CheckBox("Create Index Sheet", self.settings_card)
        self.chk_index.setChecked(True)
        
        v_opts.addWidget(self.chk_subfolders)
        v_opts.addWidget(self.chk_skip_temp)
        v_opts.addWidget(self.chk_preserve)
        v_opts.addWidget(self.chk_index)
        v_opts.addStretch()
        
        h_settings.addLayout(v_opts)
        
        # Right: Output
        v_out = QVBoxLayout()
        v_out.addWidget(StrongBodyLabel("Output", self.settings_card))
        
        h_out_path = QHBoxLayout()
        self.out_path_edit = LineEdit(self.settings_card)
        self.out_path_edit.setPlaceholderText("Output folder...")
        self.btn_out_browse = PushButton("...", self.settings_card)
        self.btn_out_browse.setFixedWidth(40)
        self.btn_out_browse.clicked.connect(self.browse_output)
        
        h_out_path.addWidget(self.out_path_edit)
        h_out_path.addWidget(self.btn_out_browse)
        
        self.out_filename_edit = LineEdit(self.settings_card)
        self.out_filename_edit.setText("MergedWorkbook.xlsx")
        self.out_filename_edit.setPlaceholderText("Filename.xlsx")
        
        self.chk_auto_open = SwitchButton("Open file after merge", self.settings_card)
        self.chk_auto_open.setChecked(True)
        
        v_out.addLayout(h_out_path)
        v_out.addWidget(self.out_filename_edit)
        v_out.addWidget(self.chk_auto_open)
        v_out.addStretch()
        
        h_settings.addLayout(v_out)
        
        settings_layout.addLayout(h_settings)
        self.v_layout.addWidget(self.settings_card)

        # Merge Button & Progress
        self.btn_merge = PrimaryPushButton("🚀 MERGE EXCEL FILES", self)
        self.btn_merge.setFixedHeight(50)
        self.btn_merge.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        self.btn_merge.clicked.connect(self.start_merge)
        self.v_layout.addWidget(self.btn_merge)

        self.progress_bar = ProgressBar(self)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.v_layout.addWidget(self.progress_bar)

        # Log
        self.log_area = TextEdit(self)
        self.log_area.setReadOnly(True)
        self.log_area.setFixedHeight(150)
        self.log_area.setPlaceholderText("Log output will appear here...")
        self.v_layout.addWidget(self.log_area)

    def browse_source(self):
        path = QFileDialog.getExistingDirectory(self, "Select Source Folder")
        if path:
            self.source_path_edit.setText(path)
            self.current_source_folder = path
            if not self.out_path_edit.text():
                self.out_path_edit.setText(path)
            self.scan_folder()

    def browse_output(self):
        path = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if path:
            self.out_path_edit.setText(path)

    def scan_folder(self):
        folder = self.source_path_edit.text()
        if not folder:
            return

        self.table.setRowCount(0)
        self.files_data = FolderScanner.scan(
            folder, 
            include_subfolders=self.chk_subfolders.isChecked(),
            skip_temp=self.chk_skip_temp.isChecked()
        )
        
        self.lbl_file_count.setText(f"{len(self.files_data)} files found")
        
        for i, info in enumerate(self.files_data):
            self.table.insertRow(i)
            
            # Checkbox item
            item_chk = QTableWidgetItem()
            item_chk.setCheckState(Qt.CheckState.Checked if info.selected else Qt.CheckState.Unchecked)
            self.table.setItem(i, 0, item_chk)
            
            self.table.setItem(i, 1, QTableWidgetItem(info.display_name))
            self.table.setItem(i, 2, QTableWidgetItem(str(info.sheet_count)))
            
            size_mb = info.path.stat().st_size / (1024 * 1024)
            self.table.setItem(i, 3, QTableWidgetItem(f"{size_mb:.2f} MB"))
            
            self.table.setItem(i, 4, QTableWidgetItem(str(info.path)))

        self.table.itemChanged.connect(self.on_item_changed)

    def on_item_changed(self, item):
        if item.column() == 0:
            row = item.row()
            if row < len(self.files_data):
                self.files_data[row].selected = (item.checkState() == Qt.CheckState.Checked)

    def toggle_all(self, select):
        state = Qt.CheckState.Checked if select else Qt.CheckState.Unchecked
        self.table.blockSignals(True)
        for i in range(self.table.rowCount()):
            self.table.item(i, 0).setCheckState(state)
            if i < len(self.files_data):
                self.files_data[i].selected = select
        self.table.blockSignals(False)

    def start_merge(self):
        if not self.files_data:
            InfoBar.warning("No files", "Please scan a folder first.", parent=self)
            return

        selected = [f for f in self.files_data if f.selected]
        if not selected:
            InfoBar.warning("No selection", "Please select at least one file to merge.", parent=self)
            return

        out_folder = self.out_path_edit.text()
        if not out_folder:
            InfoBar.error("Missing Output", "Please select an output folder.", parent=self)
            return

        settings = MergeSettings()
        settings.include_subfolders = self.chk_subfolders.isChecked()
        settings.skip_temp_files = self.chk_skip_temp.isChecked()
        settings.output_folder = pathlib.Path(out_folder)
        settings.output_filename = self.out_filename_edit.text()
        settings.create_index_sheet = self.chk_index.isChecked()
        settings.preserve_formulas = self.chk_preserve.isChecked()

        self.btn_merge.setEnabled(False)
        self.progress_bar.setValue(0)
        self.log_area.clear()
        
        self.worker = MergeWorker(self.files_data, settings)
        self.worker.log_signal.connect(self.append_log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.on_merge_finished)
        self.worker.error_signal.connect(self.on_merge_error)
        self.worker.start()

    def append_log(self, msg):
        self.log_area.append(msg)

    def update_progress(self, current, total):
        if total > 0:
            val = int((current / total) * 100)
            self.progress_bar.setValue(val)

    def on_merge_finished(self, output_path):
        self.btn_merge.setEnabled(True)
        self.progress_bar.setValue(100)
        self.last_output_path = pathlib.Path(output_path)
        
        InfoBar.success("Success", f"Merged {len([f for f in self.files_data if f.selected])} files successfully!", parent=self)
        
        if self.chk_auto_open.isChecked():
            self.open_file(output_path)

    def on_merge_error(self, err_msg):
        self.btn_merge.setEnabled(True)
        if "Permission denied" in err_msg:
            InfoBar.error(
                "File Open Error", 
                "Could not save the file. Please close 'MergedWorkbook.xlsx' and try again.", 
                parent=self,
                duration=5000
            )
        else:
            InfoBar.error("Error", f"Merge failed: {err_msg}", parent=self)
        self.log_area.append(f"CRITICAL ERROR: {err_msg}")

    def open_file(self, filepath):
        try:
            filepath_str = str(filepath)
            os.startfile(filepath_str)
        except Exception as e:
            self.append_log(f"Could not open file: {e}")

if __name__ == "__main__":
    # app is already created at module level
    window = ExcelMergerWindow()
    window.show()
    sys.exit(app.exec())
